import pandas as pd
import numpy as np
import os
from openpyxl.styles import PatternFill

# ============================================================
# --- 1. 用户自定义参数区 ---
# ============================================================

INPUT_PATH = r"C:\Users\Takumi\Desktop\临时文件\protein_galaxy\rsa_ph4.0_less.csv"
OUTPUT_PATH = r"rsa_ph4.0_less_Analysis.xlsx"
SHEET_NAME = 0

# 关键列名 (请确保与Excel一致)
COL_ID = "Protein_ID"       # 蛋白ID列
COL_AA = "Res_Name"      # 氨基酸列
COL_RSA = "RSA" # <--- 修改点1: 请确保这里填写的名字和你Excel表头完全一样

# 20种标准氨基酸
STD_AAS = sorted(['A', 'R', 'N', 'D', 'C', 'E', 'Q', 'G', 'H', 'I', 
                  'L', 'K', 'M', 'F', 'P', 'S', 'T', 'W', 'Y', 'V'])

# 梯度颜色定义 (Rank 1 -> Rank 5)
RANK_COLORS = [
    "E06666", # Rank 1 (最深)
    "EA9999", # Rank 2
    "F4CCCC", # Rank 3
    "FCE5CD", # Rank 4
    "FFF2CC"  # Rank 5 (最浅)
]

# ============================================================
# --- 2. 核心逻辑 ---
# ============================================================

def apply_gradient_highlight(ws):
    """
    对工作表的每一列数据应用 Top 5 梯度着色
    """
    for col in ws.iter_cols(min_col=2, min_row=2, max_row=ws.max_row):
        values = []
        cells = []
        for cell in col:
            if isinstance(cell.value, (int, float)):
                values.append(cell.value)
                cells.append(cell)
        
        if not values:
            continue
            
        unique_values = sorted(list(set(values)), reverse=True)
        top5_values = unique_values[:5]
        
        value_color_map = {}
        for i, val in enumerate(top5_values):
            if i < len(RANK_COLORS):
                color_hex = RANK_COLORS[i]
                value_color_map[val] = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
        
        for cell in cells:
            if cell.value in value_color_map:
                cell.fill = value_color_map[cell.value]

def main():
    print(f"--- 开始处理 (V3 修正版 - 找回 Total 列) ---")
    
    # --- 1. 读取数据 ---
    if not os.path.exists(INPUT_PATH):
        print(f"错误: 文件不存在 {INPUT_PATH}")
        return

    try:
        if INPUT_PATH.endswith('.csv'):
            try: df = pd.read_csv(INPUT_PATH, on_bad_lines='skip')
            except: df = pd.read_csv(INPUT_PATH, error_bad_lines=False)
        else:
            df = pd.read_excel(INPUT_PATH, sheet_name=SHEET_NAME)
        print(f"读取成功，共 {len(df)} 行。")
    except Exception as e:
        print(f"读取失败: {e}")
        return

    # 检查列
    req_cols = [COL_ID, COL_AA, COL_RSA]
    if not all(c in df.columns for c in req_cols):
        print(f"错误: 缺列 {req_cols}")
        return

    # --- 2. 清洗 ---
    df[COL_RSA] = pd.to_numeric(df[COL_RSA], errors='coerce')
    
    if df[COL_RSA].max() > 1.0:
        print("检测到百分比数值，正在转换为 0-1 小数...")
        df[COL_RSA] = df[COL_RSA] / 100.0

    mask = (df[COL_RSA] >= 0) & (df[COL_RSA] <= 1)
    df_clean = df[mask].copy()
    df_invalid = df[~mask].copy()
    
    df_clean[COL_AA] = df_clean[COL_AA].astype(str).str.strip().str.upper()

    bins = [0, 0.2, 0.4, 0.6, 0.8, 1.0]
    labels = ['0-0.2', '0.2-0.4', '0.4-0.6', '0.6-0.8', '0.8-1']
    df_clean['RSA_Bin'] = pd.cut(df_clean[COL_RSA], bins=bins, labels=labels, include_lowest=True)

    # ========================================================
    # Task 1: 计数 (Table 1)
    # ========================================================
    print("生成表 1 (Counts)...")
    table_counts = df_clean.pivot_table(index=COL_AA, columns='RSA_Bin', aggfunc='size', fill_value=0, observed=False)
    
    # ========================================================
    # Task 2: 频率 (Table 2) - 先计算这个，再改 Table 1
    # ========================================================
    print("生成表 2 (Range Composition)...")
    col_sums = table_counts.sum(axis=0) 
    table_freq = table_counts.div(col_sums, axis=1).fillna(0).round(4)

    # --- [修复点] 在 Task 2 计算完成后，给 Table 1 补上 Total 列 ---
    table_counts['Total'] = table_counts.sum(axis=1)
    # -----------------------------------------------------------

    # ========================================================
    # Task 3: Top 5 per Protein (Table 3)
    # ========================================================
    print("生成表 3 (Protein Top 5)...")
    grp = df_clean.groupby([COL_ID, COL_AA]).size().reset_index(name='count')
    totals = df_clean.groupby(COL_ID).size().reset_index(name='total')
    grp = pd.merge(grp, totals, on=COL_ID)
    grp['per'] = (grp['count'] / grp['total']).round(4)
    grp_sorted = grp.sort_values(by=[COL_ID, 'per'], ascending=[True, False])

    def get_top5_row(sub_df):
        top = sub_df.head(5)
        res = {}
        for i, (idx, row) in enumerate(top.iterrows()):
            n = i + 1
            res[f't{n}_aa'] = row[COL_AA]
            res[f't{n}_per'] = row['per']
        return pd.Series(res)

    task3_res = grp_sorted.groupby(COL_ID).apply(get_top5_row, include_groups=False).reset_index()
    
    expected_cols = []
    for i in range(1, 6):
        expected_cols.extend([f't{i}_aa', f't{i}_per'])
    
    for c in expected_cols:
        if c not in task3_res.columns: task3_res[c] = None
    task3_res = task3_res[[COL_ID] + expected_cols]

    # ========================================================
    # Task 4: All Ranges Dominant (Table 4)
    # ========================================================
    print("生成表 4 (All Ranges Dominant)...")
    all_ids = pd.DataFrame(df_clean[COL_ID].unique(), columns=[COL_ID])
    
    def get_dom_aa(bin_label):
        sub = df_clean[df_clean['RSA_Bin'] == bin_label]
        if sub.empty: return pd.DataFrame(columns=[COL_ID, f'{bin_label}_aa'])
        
        cnt = sub.groupby([COL_ID, COL_AA]).size().reset_index(name='c')
        top = cnt.sort_values([COL_ID, 'c'], ascending=[True, False]).drop_duplicates(COL_ID)
        return top[[COL_ID, COL_AA]].rename(columns={COL_AA: f'{bin_label}_aa'})

    task4_res = all_ids
    for label in labels:
        res_bin = get_dom_aa(label)
        task4_res = pd.merge(task4_res, res_bin, on=COL_ID, how='left')
    
    task4_res = task4_res.fillna('None')

    # ========================================================
    # Task 5: Stats of Top 1-5 (Table 5)
    # ========================================================
    print("生成表 5 (Stats Top 1-5)...")
    total_prots = len(task3_res)
    task5_df = pd.DataFrame(index=STD_AAS)
    
    def calc_freq(series):
        c = series.value_counts().reindex(STD_AAS, fill_value=0)
        return (c / total_prots).round(4)

    for i in range(1, 6):
        col_name = f't{i}_aa'
        if col_name in task3_res.columns:
            task5_df[f'Freq_in_T{i}'] = calc_freq(task3_res[col_name])
        else:
            task5_df[f'Freq_in_T{i}'] = 0.0

    task5_res = task5_df.reset_index().rename(columns={'index': 'AA'})

    # ========================================================
    # Task 6: Stats of All Ranges (Table 6)
    # ========================================================
    print("生成表 6 (Stats Ranges)...")
    task6_df = pd.DataFrame(index=STD_AAS)
    
    def calc_freq_rsa(series):
        c = series.value_counts().reindex(STD_AAS, fill_value=0)
        return (c / len(task4_res)).round(4)

    for label in labels:
        col_name = f'{label}_aa'
        if col_name in task4_res.columns:
            task6_df[f'Freq_{label}'] = calc_freq_rsa(task4_res[col_name])
    
    task6_res = task6_df.reset_index().rename(columns={'index': 'AA'})

    # ========================================================
    # 保存与着色
    # ========================================================
    print(f"正在保存并应用梯度着色: {OUTPUT_PATH}")
    
    try:
        with pd.ExcelWriter(OUTPUT_PATH, engine='openpyxl') as writer:
            table_counts.to_excel(writer, sheet_name='1.RSA_Counts')
            table_freq.to_excel(writer, sheet_name='2.AA_Composition_in_Range')
            task3_res.to_excel(writer, sheet_name='3.Protein_Top5', index=False)
            task4_res.to_excel(writer, sheet_name='4.Ranges_Dominant', index=False)
            task5_res.to_excel(writer, sheet_name='5.Stats_Top5_Freq', index=False)
            task6_res.to_excel(writer, sheet_name='6.Stats_Ranges_Freq', index=False)
            
            if not df_invalid.empty:
                df_invalid.to_excel(writer, sheet_name='Error_Log', index=False)
            
            apply_gradient_highlight(writer.sheets['1.RSA_Counts'])
            apply_gradient_highlight(writer.sheets['2.AA_Composition_in_Range'])
            apply_gradient_highlight(writer.sheets['5.Stats_Top5_Freq'])
            apply_gradient_highlight(writer.sheets['6.Stats_Ranges_Freq'])
            
        print("--- 完成! ---")
        
    except Exception as e:
        print(f"保存失败: {e}")

if __name__ == "__main__":
    main()